#coding=utf-8
import pytest
import create_excel
import simple_module
import json
import sys
import os
from ansible.module_utils._text import to_bytes
from ansible.module_utils import basic
from ansible.compat.tests import unittest
from cStringIO import StringIO

from openpyxl import Workbook, cell
from openpyxl.styles import Font, PatternFill, Border, Protection, Alignment, Side

# Test data for create_excel
# Basic data (should create new files with first call and do nothing with second call)
EXCEL_DATA_1 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D1','E1','D2','E2'],['D3','E3','D4','E4'],['D5','E5','D6','E6'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2.xlsx',
}

# Changed basic data with the same output files (should rewrite that file if called after basic data)
EXCEL_DATA_2 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D100','E100','D200','E200'],['D300','E300','D400','E400'],['D500','E500','D600','E600'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2.xlsx',
}

# Data with changed output files (should write new files)
EXCEL_DATA_3 = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D100','E100','D200','E200'],['D300','E300','D400','E400'],['D500','E500','D600','E600'],],
                    'input_excel':'/syncfolder/excel_test_files/input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1_new.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2_new.xlsx',
}

# Data with no input file - should fail module with corresponding message
NO_INPUT_FILE_EXCEL_DATA = {'header_data':['Name A', 'Interface A', 'Name B', 'Interface B'],
                    'function_name':'full copy',
                    'table_data':[['D1','E1','D2','E2'],['D3','E3','D4','E4'],['D5','E5','D6','E6'],],
                    'input_excel':'/syncfolder/no_excel_test_files/no_input_excel_uri.xlsx',
                    'output_excel1':'/syncfolder/excel_test_files/output_excel_uri1_new.xlsx',
                    'output_excel2':'/syncfolder/excel_test_files/output_excel_uri2_new.xlsx',
}

#
# Data for simple module tests
#
SIMPLE_DATA_1 = {'name':'Alex1',
                 'secret':'secret1',}
SIMPLE_DATA_2 = {'name':'Alex2',
                 'secret':'secret2',}
SIMPLE_DATA_3 = {'name':'Alex3',
                 'secret':'secret3',}

#
# Function sets module arguments
#
def set_module_args(args):
    args = json.dumps({"ANSIBLE_MODULE_ARGS": args})
    basic._ANSIBLE_ARGS = to_bytes(args)

#
# link1: https://docs.pytest.org/en/latest/xunit_setup.html for setUp and tearDown
# link2: https://stackoverflow.com/questions/26405380/how-do-i-correctly-setup-and-teardown-my-pytest-class-with-tests
# This is used to replace setUp and tearDown functions for TestExcelModule class
@pytest.fixture()
def resource():
    backup = sys.stdout
    sys.stdout = StringIO()
    yield "resource"
    sys.stdout.close()
    sys.stdout = backup

#
# This class is used to test simple_module and create_excel modules (1st variant, the best IMHO)
#
class TestExcelModule(object):
    
    #
    # simple_module test with different input parameters
    #
    @pytest.mark.parametrize('test_input,expected', [
        (SIMPLE_DATA_1, 'Successfully called simple module'),
        (SIMPLE_DATA_2, 'Successfully called simple module'),
        (SIMPLE_DATA_3, 'Successfully called simple module'),
    ])
    def test_simple_module_has_correct_output(self, test_input, expected, resource):
        # setting ansible arguments on stdin
        set_module_args(test_input)
        with pytest.raises(SystemExit):
            simple_module.main()
        # capturing output in console 
        output = sys.stdout.getvalue()
        # assertion statement checks if output corresponds to expected
        assert expected in output
    
    #
    # create_excel test with different input parameters
    #
    @pytest.mark.parametrize('test_input,expected', [
        (EXCEL_DATA_1, 'Successfully copied excel data'),
        (EXCEL_DATA_1, 'File already exists and correct. Do nothing'),
        (EXCEL_DATA_2, 'Successfully copied excel data'),
        (EXCEL_DATA_3, 'Successfully copied excel data'),
        (NO_INPUT_FILE_EXCEL_DATA, 'File cannot be open'),
    ])    
    def test_create_excel_has_correct_output(self, test_input, expected, resource):
        # This is used to clear the directory from output files to be able to reuse test
        if (test_input == NO_INPUT_FILE_EXCEL_DATA):
            if os.path.isfile(EXCEL_DATA_1['output_excel1']):
                os.remove(EXCEL_DATA_1['output_excel1'])
            if os.path.isfile(EXCEL_DATA_3['output_excel1']):
                os.remove(EXCEL_DATA_3['output_excel1'])
        set_module_args(test_input)
        with pytest.raises(SystemExit):
            create_excel.main()
        output = sys.stdout.getvalue()
        assert expected in output

#
# This is the version of simple_module test outside of class
# Have to set the stdout capture every time it is called
# Less preferable
#
@pytest.mark.parametrize('test_input,expected', [
        (SIMPLE_DATA_1, 'Successfully called simple module'),
        (SIMPLE_DATA_2, 'Successfully called simple module'),
        (SIMPLE_DATA_3, 'Successfully called simple module'),
])
def test_module_has_correct_output(test_input, expected):
    # here start reading stdout
    backup = sys.stdout
    sys.stdout = StringIO()
    # setting ansible arguments on stdin
    set_module_args(test_input)
    with pytest.raises(SystemExit):
        simple_module.main()
    output = sys.stdout.getvalue()
    sys.stdout.close()
    sys.stdout = backup
    assert expected in output
    
#
# This is another version of class that tests simple_module
# TestCase is used, so we can define setUp and tearDown methods inside 
# Cannot use parameters in that case - have to call method multiple times to check different cases 
#
class TestExcelModuleWithUnittest(unittest.TestCase):
    
    def setUp(self):
        # here start reading stdout
        self.backup = sys.stdout
        sys.stdout = StringIO()
 
    def tearDown(self):
        # here stop reading stdout
        sys.stdout.close()
        sys.stdout = self.backup
    #
    # Function tests output of the module
    # Unfortunately cannot use parameters in that case, so we use only one test case there
    #
    def test_module_has_correct_output(self):
        # setting ansible arguments on stdin
        set_module_args({'name':'Alex1',
                         'secret':'secret1',})
        with self.assertRaises(SystemExit) as result:
            simple_module.main()
        # For Python 2.6 compatibility
        if isinstance(result.exception, int):
            self.assertEquals(result.exception, 0)
        else:
            self.assertEquals(result.exception.code, 0)
        output = sys.stdout.getvalue()
        assert 'Successfully called simple module' in output
        
    #
    # In that case function is called once, so we do not need multiple parameters
    # Function tests if module fails if there is no input arguments
    #    
    def test_module_fails_when_required_args_missing(self):
        set_module_args({})
        with self.assertRaises(SystemExit) as result:
            simple_module.run_module()
        if isinstance(result.exception, int):
            self.assertEquals(result.exception, 1)
        else:
            self.assertEquals(result.exception.code, 1)
        output = sys.stdout.getvalue()
        assert 'missing required arguments: secret, name' in output
            

#
# Another variant does not work (from official documentation):
# https://docs.ansible.com/ansible/2.5/dev_guide/testing_units_modules.html        
# 
 
#
# Test compare_style function
# https://openpyxl.readthedocs.io/en/stable/_modules/openpyxl/styles/borders.html
# This is simple unit test of internal module function that compares styles of excel cells 
#
def test_styles_are_equal_function():
    
    new_workbook = Workbook()
    new_workbook.create_sheet("Sheet1", 0)
    new_worksheet1 = new_workbook["Sheet1"]
    new_font = Font(name='Calibri',
                    size=11,
                    bold=False,
                    italic=True,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')
    new_fill = PatternFill(fill_type=None,
                    start_color='FFFFFFFF',
                    end_color='FF000000')
    
    new_border = Border(left=Side(border_style=None,
                                  color='FF000000'),
                        right=Side(border_style=None,
                                   color='FF000000'),
                        top=Side(border_style=None,
                                 color='FF000000'),
                        bottom=Side(border_style=None,
                                    color='FF000000'),
                        diagonal=Side(border_style=None,
                                      color='FF000000'),
                        diagonal_direction=0,
                        outline=True,
                        vertical=Side(border_style=None,
                                      color='FF000000'),
                        horizontal=Side(border_style=None,
                                        color='FF000000')

                        )
    
    new_alignment=Alignment(horizontal='general',
                            vertical='bottom',
                            text_rotation=0,
                            wrap_text=False,
                            shrink_to_fit=False,
                            indent=0)
    
    new_number_format = 'General'
    
    new_protection = Protection(locked=True,
                                hidden=False)
    
    cell1 = new_worksheet1.cell(row = 1, column = 1)
    cell1.font = new_font
    cell1.fill = new_fill
    cell1.border = new_border
    cell1.protection = new_protection
    cell1.alignment = new_alignment
    cell1.number_format = new_number_format
    
    cell2 = new_worksheet1.cell(row = 1, column = 2)
    cell2.font = new_font
    cell2.fill = new_fill
    cell2.border = new_border
    cell2.protection = new_protection
    cell2.alignment = new_alignment
    cell2.number_format = new_number_format
    
    assert create_excel.areStylesEqual(cell1, cell2) == True
