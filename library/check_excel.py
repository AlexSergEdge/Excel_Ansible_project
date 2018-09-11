#coding=utf-8
from ansible.module_utils.basic import *
from openpyxl import Workbook, cell
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, Fill, Border, Protection, Alignment
import os.path
from openpyxl.styles.builtins import output

#
# Function is used to compare styles of two cells 
# https://openpyxl.readthedocs.io/en/2.5/styles.html
#
def areStylesEqual(toStyle, fromStyle):
    return (toStyle.font.name == fromStyle.font.name
            and toStyle.font.size == fromStyle.font.size
            and toStyle.font.bold == fromStyle.font.bold
            and toStyle.font.italic == fromStyle.font.italic
            and toStyle.font.vertAlign == fromStyle.font.vertAlign
            and toStyle.font.underline == fromStyle.font.underline
            and toStyle.font.strike == fromStyle.font.strike
            and toStyle.font.color == fromStyle.font.color
            
            and toStyle.fill.fill_type == fromStyle.fill.fill_type
            and toStyle.fill.start_color == fromStyle.fill.start_color
            and toStyle.fill.end_color == fromStyle.fill.end_color
            
            and toStyle.border.left.border_style == fromStyle.border.left.border_style
            and toStyle.border.left.color == fromStyle.border.left.color
            and toStyle.border.right.border_style == fromStyle.border.right.border_style
            and toStyle.border.right.color == fromStyle.border.right.color
            and toStyle.border.top.border_style == fromStyle.border.top.border_style
            and toStyle.border.top.color == fromStyle.border.top.color
            and toStyle.border.bottom.border_style == fromStyle.border.bottom.border_style
            and toStyle.border.bottom.color == fromStyle.border.bottom.color
            and toStyle.border.diagonal.border_style == fromStyle.border.diagonal.border_style
            and toStyle.border.diagonal.color == fromStyle.border.diagonal.color
            and toStyle.border.diagonal_direction == fromStyle.border.diagonal_direction
            and toStyle.border.outline == fromStyle.border.outline
            
            # This attributes are not copied by copyStyle function
            
            # and toStyle.border.vertical.border_style == fromStyle.border.vertical.border_style
            # and toStyle.border.vertical.color == fromStyle.border.vertical.color
            
            # and toStyle.border.horizontal.border_style == fromStyle.border.horizontal.border_style
            # and toStyle.border.horizontal.color == fromStyle.border.horizontal.color
             
            and toStyle.alignment.horizontal == fromStyle.alignment.horizontal
            and toStyle.alignment.vertical == fromStyle.alignment.vertical
            and toStyle.alignment.text_rotation == fromStyle.alignment.text_rotation
            and toStyle.alignment.wrap_text == fromStyle.alignment.wrap_text
            and toStyle.alignment.shrink_to_fit == fromStyle.alignment.shrink_to_fit
            and toStyle.alignment.indent == fromStyle.alignment.indent
            
            and toStyle.number_format == fromStyle.number_format
            
            and toStyle.protection.locked == fromStyle.protection.locked
            and toStyle.protection.hidden == fromStyle.protection.hidden)

# Function compares headers of input and output file
def compare_header(input_excel, output_excel):
    input_workbook = Workbook()
    output_workbook = Workbook()
    input_workbook = load_workbook(input_excel)    
    output_workbook = load_workbook(output_excel)
    input_worksheet = input_workbook["Sheet1"]
    output_worksheet = output_workbook["Sheet1"]
    comparsion_failed = False
    real_header_length = 0
    # Counting number of headers in input file
    for col in input_worksheet.iter_cols():
        for cell in col:
                real_header_length = real_header_length + 1
    for i in range(1, real_header_length+1):
        outputCell = output_worksheet.cell(row = 1, column = i)
        inputCell = input_worksheet.cell(row = 1, column = i)
        # If any output cell value is not equal to input cell value - set fail True
        if (outputCell.value != inputCell.value) or (areStylesEqual(outputCell,inputCell) == False):
            comparsion_failed = True      
    return comparsion_failed, False, real_header_length

# Function is used to compare data in output file and table_data
def check_data(header_data, table_data, output_excel, real_header_length):
    output_workbook = Workbook()
    data_test_failed = False
    output_workbook = load_workbook(output_excel)
    output_worksheet = output_workbook["Sheet1"]
    # Looking for data where header name is equal to the name from header_data list 
    for i in range(0, len(table_data[0])):
        for j in range(0, real_header_length):
            if header_data[i] == output_worksheet.cell(row=1, column=j+1).value:
                for k in range(0, len(table_data)):
                    # If any output data is not equal to input data, set fail True
                    if output_worksheet.cell(row=k+2, column=j+1).value != table_data[k][i]:
                        data_test_failed = True
    return data_test_failed, False

def run_module():
    fields = dict(
        header_data = dict(required=True, type='list'),
        function_name = dict(default="full copy", choises=["full copy","workbook copy"], type='str'), 
        table_data = dict(required=True, type='list'),
        input_excel = dict(required=True, type='str'),
        output_excel1 = dict(required=True, type='str'),
        output_excel2 = dict(required=True, type='str'),
    )
    result = dict(
        changed=False,
        failed=False,
        message=''
    )
    module = AnsibleModule(argument_spec=fields, supports_check_mode=True)
    if module.check_mode:
        return result

    # get all input data
    header_data = module.params['header_data']
    function_name = module.params['function_name']
    table_data = module.params['table_data']
    input_excel = module.params['input_excel']
    output_excel1 = module.params['output_excel1']
    output_excel2 = module.params['output_excel2']
    
    
    try:
    # First variant
        if function_name == "full copy":
            comparsion_failed, header_changed, real_header_length = compare_header(input_excel, output_excel1)
            data_test_failed, data_changed = check_data(header_data, table_data, output_excel1, real_header_length)
            result['failed'] = comparsion_failed or data_test_failed
            result['changed'] = data_changed or header_changed
            result['message'] = "Successfully checked excel data"
            module.exit_json(**result)
        # If second variant is selected (using same functions, but pass output_excel2 file in them)
        elif function_name == "workbook copy":
            comparsion_failed, header_changed, real_header_length = compare_header(input_excel, output_excel2)
            data_test_failed, data_changed = check_data(header_data, table_data, output_excel2, real_header_length)
            result['failed'] = comparsion_failed or data_test_failed
            result['changed'] = data_changed or header_changed
            result['message'] = "Successfully checked excel data"
            module.exit_json(**result)
    except IOError:
        module.exit_json(changed=False, failed=True, meta="File cannot be open")
        
def main():
    run_module()

if __name__ == '__main__':
    main()

