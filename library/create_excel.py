#coding=utf-8
from ansible.module_utils.basic import *
from openpyxl import Workbook, cell
from openpyxl import load_workbook
from copy import copy
from openpyxl.styles import Font, PatternFill, Border, Protection, Alignment, Side
import os.path

#
# Function is used in copy_header function to copy full cell style
#
def copyStyle(toStyle, fromStyle):
    toStyle.font = copy(fromStyle.font)
    toStyle.fill = copy(fromStyle.fill)
    toStyle.border = copy(fromStyle.border)
    toStyle.alignment = copy(fromStyle.alignment)
    toStyle.number_format = copy(fromStyle.number_format)
    toStyle.protection = copy(fromStyle.protection)

#
# Function is used to compare styles of two cells 
# https://openpyxl.readthedocs.io/en/2.5/styles.html
#
def areStylesEqual(toStyle, fromStyle):
    return (toStyle.font.name == fromStyle.font.name
            and toStyle.font.size == fromStyle.font.size
            and toStyle.font.bold == fromStyle.font.bold
            and toStyle.font.italic == fromStyle.font.italic
            and toStyle.font.vertAlign == fromStyle.font.vertAlign
            and toStyle.font.underline == fromStyle.font.underline
            and toStyle.font.strike == fromStyle.font.strike
            and toStyle.font.color == fromStyle.font.color
            
            and toStyle.fill.fill_type == fromStyle.fill.fill_type
            and toStyle.fill.start_color == fromStyle.fill.start_color
            and toStyle.fill.end_color == fromStyle.fill.end_color
            
            and toStyle.border.left.border_style == fromStyle.border.left.border_style
            and toStyle.border.left.color == fromStyle.border.left.color
            and toStyle.border.right.border_style == fromStyle.border.right.border_style
            and toStyle.border.right.color == fromStyle.border.right.color
            and toStyle.border.top.border_style == fromStyle.border.top.border_style
            and toStyle.border.top.color == fromStyle.border.top.color
            and toStyle.border.bottom.border_style == fromStyle.border.bottom.border_style
            and toStyle.border.bottom.color == fromStyle.border.bottom.color
            and toStyle.border.diagonal.border_style == fromStyle.border.diagonal.border_style
            and toStyle.border.diagonal.color == fromStyle.border.diagonal.color
            and toStyle.border.diagonal_direction == fromStyle.border.diagonal_direction
            and toStyle.border.outline == fromStyle.border.outline
            
            # This attributes are not copied by copyStyle function
            
            # and toStyle.border.vertical.border_style == fromStyle.border.vertical.border_style
            # and toStyle.border.vertical.color == fromStyle.border.vertical.color
            
            # and toStyle.border.horizontal.border_style == fromStyle.border.horizontal.border_style
            # and toStyle.border.horizontal.color == fromStyle.border.horizontal.color
             
            and toStyle.alignment.horizontal == fromStyle.alignment.horizontal
            and toStyle.alignment.vertical == fromStyle.alignment.vertical
            and toStyle.alignment.text_rotation == fromStyle.alignment.text_rotation
            and toStyle.alignment.wrap_text == fromStyle.alignment.wrap_text
            and toStyle.alignment.shrink_to_fit == fromStyle.alignment.shrink_to_fit
            and toStyle.alignment.indent == fromStyle.alignment.indent
            
            and toStyle.number_format == fromStyle.number_format
            
            and toStyle.protection.locked == fromStyle.protection.locked
            and toStyle.protection.hidden == fromStyle.protection.hidden)

# Function takes input file, copies header cells in list, and then creates new output file (or takes existing) and copies cells in it
def copy_header(input_excel, output_excel):
    
    input_workbook = Workbook()
    output_workbook = Workbook()    
    input_workbook = load_workbook(input_excel)
    # Get sheets to work with
    input_worksheet = input_workbook["Sheet1"]
    # flag says if file has been modified
    has_changed = False
    # Copying header cell into the list
    header_values = []
    for col in input_worksheet.iter_cols():
        for cell in col:
                header_values.append(cell.value)
    # Check if file already exists
    if (os.path.isfile(output_excel)):
        # Checking if header exists and is correct
        output_workbook = load_workbook(output_excel)
        output_worksheet = output_workbook["Sheet1"] # get_sheet_by_name('default')
        for i in range(1, len(header_values)+1):
            outputCell = output_worksheet.cell(row = 1, column = i)
            inputCell = input_worksheet.cell(row = 1, column = i)
            # If something is wrong, it corrects this file 
            if (outputCell.value != header_values[i - 1]) or not areStylesEqual(outputCell,inputCell):
                outputCell.value = header_values[i - 1]
                copyStyle(outputCell,inputCell)
                has_changed = True
        # Save all changes made to file
        output_workbook.save(output_excel)
    # If there is no output file
    else:
        has_changed = True
        # Creating new worksheet in output_workbook
        output_workbook.create_sheet("Sheet1", 0)
        output_worksheet = output_workbook["Sheet1"]
        # Inserting all header data into the new file
        for i in range(1, len(header_values)+1):
            outputCell = output_worksheet.cell(row = 1, column = i)
            inputCell = input_worksheet.cell(row = 1, column = i)
            outputCell.value = header_values[i - 1]
            copyStyle(outputCell,inputCell)
        # Saving output file
        output_workbook.save(output_excel)
    return False, has_changed, len(header_values)


# Function is used to copy data from list of data to the output file
def copy_data(header_data, table_data, output_excel, real_header_length):

    output_workbook = Workbook()
    data_has_changed = False
    output_workbook = load_workbook(output_excel)
    output_worksheet = output_workbook["Sheet1"]
    # Iterating through cells and table_data 
    for i in range(0, len(table_data[0])):
        for j in range(0, real_header_length):
            # Inserting data where header name is equal to the name from header_data list
            if header_data[i] == output_worksheet.cell(row=1, column=j+1).value:
                for k in range(0, len(table_data)):
                    # Check if data is already exists and correct
                    if output_worksheet.cell(row=k+2, column=j+1).value != table_data[k][i]:
                        output_worksheet.cell(row=k+2, column=j+1).value = table_data[k][i]
                        data_has_changed = True
    # Save all changes made to file
    output_workbook.save(output_excel)
    return False, data_has_changed


# Function saves input file in memory, fills it with data, and saves it as a new output file
def create_excel(header_data, input_excel, output_excel, table_data):

    has_changed = True
    # Creating workbook
    input_workbook = Workbook()
    # Opening file
    input_workbook = load_workbook(input_excel)
    # Select work sheet
    input_worksheet = input_workbook["Sheet1"]
    # Real header size
    header_size=0
    for col in input_worksheet.iter_cols():
        for cell in col:
            if cell.value != '':
                header_size += 1
    # filter map . . . 
    # Filling cell with data
    for i in range(0, len(table_data[0])):
        for j in range(0, header_size):
            if header_data[i] == input_worksheet.cell(row=1, column=j+1).value:
                for k in range(0, len(table_data)):
                    input_worksheet.cell(row=k+2, column=j+1).value = table_data[k][i]
    # Saving as a new file
    input_workbook.save(output_excel)
    # return flags Failed, Changed
    return False, has_changed


def run_module():
    fields = dict(
        header_data = dict(required=True, type='list'),
        function_name = dict(default="full copy", choises=["full copy","workbook copy"], type='str'),
        table_data = dict(required=True, type='list'),
        input_excel = dict(required=True, type='str'),
        output_excel1 = dict(required=True, type='str'),
        output_excel2 = dict(required=True, type='str')   
    )
    
    result = dict(
        changed=False,
        message=''
    )
    
    module = AnsibleModule(argument_spec=fields, supports_check_mode=True)
    
    if module.check_mode:
        return result
    
    # get all input data
    header_data = module.params['header_data']
    function_name = module.params['function_name']
    table_data = module.params['table_data']
    input_excel = module.params['input_excel']
    output_excel1 = module.params['output_excel1']
    output_excel2 = module.params['output_excel2']
    
    # If first variant (create new file and copy header cell by cell in it) is selected (2 steps required)
    try:
        if function_name == "full copy":
            # Copy header
            header_failed, header_changed, real_header_length = copy_header(input_excel, output_excel1)
            data_failed, data_changed = copy_data(header_data, table_data, output_excel1, real_header_length)
            result['changed'] = data_changed or header_changed
            if result['changed']:
                result['message'] = "Successfully copied excel data"
                module.exit_json(**result)
            else:
                result['message'] = "File already exists and correct. Do nothing"
                module.exit_json(**result)
        # If second variant is selected - only one function call required
        elif function_name == "workbook copy":
            is_failed, is_changed = create_excel(header_data, input_excel, output_excel2, table_data)
            if is_changed:
                module.exit_json(changed=is_changed, failed=is_failed, meta="Successfully called create_excel function")
            else:
                module.exit_json(changed=is_changed, failed=is_failed, meta="File already exists and correct. Do nothing")
    except IOError:
        # module.fail_json(msg="File cannot be open")
        module.exit_json(changed=False, failed=True, meta="File cannot be open")

def main():
    run_module()

if __name__ == '__main__':
    main()
